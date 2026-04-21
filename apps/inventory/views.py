"""
Views for inventory app.
Handles products, categories, batches, and stock management.
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.db.models import Sum, F, Q, Avg
from django.http import JsonResponse, HttpResponse
from decimal import Decimal
import decimal

from apps.core.mixins import PaginationMixin # Added this line

from .models import Category, Product, Batch, InventoryLedger, ShopPrice, StockAdjustment
from .forms import CategoryForm, ProductForm, BatchForm, StockAdjustmentForm, ShopPriceForm
from apps.core.models import Location
from apps.core.mixins import PaginationMixin, SortableMixin

import openpyxl


# ============ Category Views ============
class CategoryListView(LoginRequiredMixin, SortableMixin, ListView):
    """List all categories for the tenant."""
    model = Category
    template_name = 'inventory/category_list.html'
    context_object_name = 'categories'
    sortable_fields = ['name', 'parent__name', 'created_at']
    default_sort = 'name'
    
    def get_queryset(self):
        queryset = Category.objects.filter(
            tenant=self.request.user.tenant
        ).select_related('parent')
        return self.apply_sorting(queryset)


class CategoryCreateView(LoginRequiredMixin, CreateView):
    """Create a new category."""
    model = Category
    form_class = CategoryForm
    template_name = 'inventory/category_form.html'
    success_url = reverse_lazy('inventory:category_list')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['tenant'] = self.request.user.tenant
        return kwargs
    
    def form_valid(self, form):
        form.instance.tenant = self.request.user.tenant
        messages.success(self.request, f'Category "{form.instance.name}" created successfully!')
        return super().form_valid(form)


class CategoryUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing category."""
    model = Category
    form_class = CategoryForm
    template_name = 'inventory/category_form.html'
    success_url = reverse_lazy('inventory:category_list')
    
    def get_queryset(self):
        return Category.objects.filter(tenant=self.request.user.tenant)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['tenant'] = self.request.user.tenant
        return kwargs
    
    def form_valid(self, form):
        messages.success(self.request, f'Category "{form.instance.name}" updated!')
        return super().form_valid(form)


class CategoryDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a category."""
    model = Category
    template_name = 'inventory/category_confirm_delete.html'
    success_url = reverse_lazy('inventory:category_list')
    
    def get_queryset(self):
        return Category.objects.filter(tenant=self.request.user.tenant)


# ============ Product Views ============
class ProductListView(LoginRequiredMixin, SortableMixin, ListView):
    """List all products for the tenant."""
    model = Product
    template_name = 'inventory/product_list.html'
    context_object_name = 'products'
    sortable_fields = ['name', 'sku', 'default_selling_price', 'total_stock', 'category__name']
    default_sort = 'name'
    
    def _get_user_location(self):
        """Get the user's effective location."""
        user = self.request.user
        if user.location:
            return user.location
        if user.role:
            role_location_map = {
                'PRODUCTION_MANAGER': 'PRODUCTION',
                'STORES_MANAGER': 'STORES',
                'SHOP_MANAGER': 'SHOP',
            }
            loc_type = role_location_map.get(user.role.name)
            if loc_type:
                return Location.objects.filter(
                    tenant=user.tenant, is_active=True, location_type=loc_type
                ).first()
        return None
    
    def get_queryset(self):
        from django.db.models import Sum, Subquery, OuterRef, Value, IntegerField, Case, When
        from django.db.models.functions import Coalesce
        from decimal import Decimal
        from .models import InventoryLedger, FavoriteProduct
        
        user = self.request.user
        queryset = Product.objects.filter(
            tenant=user.tenant
        ).select_related('category')
        
        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | 
                Q(sku__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Category filter
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)
        
        # Annotate total_stock for sorting
        user_location = user.location
        user_location_type = None
        
        if not user_location and user.role:
            role_location_map = {
                'PRODUCTION_MANAGER': 'PRODUCTION',
                'STORES_MANAGER': 'STORES',
                'SHOP_MANAGER': 'SHOP',
            }
            user_location_type = role_location_map.get(user.role.name)
            
        stock_filter = {'product': OuterRef('pk'), 'tenant': user.tenant}
        if user_location:
            stock_filter['location'] = user_location
        elif user_location_type:
            stock_filter['location__location_type'] = user_location_type
            
        stock_subquery = InventoryLedger.objects.filter(**stock_filter).order_by().values('product').annotate(
            total=Sum('quantity')
        ).values('total')
        
        queryset = queryset.annotate(
            total_stock=Coalesce(Subquery(stock_subquery), Decimal('0.00'))
        )
        
        # Annotate favorites: is_favorite = 1 for favorites, 0 for others
        effective_location = self._get_user_location()
        if effective_location:
            favorite_ids = set(FavoriteProduct.objects.filter(
                tenant=user.tenant, location=effective_location
            ).values_list('product_id', flat=True))
            
            if favorite_ids:
                queryset = queryset.annotate(
                    is_favorite=Case(
                        When(pk__in=favorite_ids, then=Value(0)),
                        default=Value(1),
                        output_field=IntegerField(),
                    )
                )
            else:
                queryset = queryset.annotate(
                    is_favorite=Value(1, output_field=IntegerField())
                )
        else:
            favorite_ids = set()
            queryset = queryset.annotate(
                is_favorite=Value(1, output_field=IntegerField())
            )
        
        # Apply sorting: favorites first, then the user's chosen sort
        sorted_qs = self.apply_sorting(queryset)
        # Prepend is_favorite to the ordering so favorites always come first
        current_ordering = sorted_qs.query.order_by
        sorted_qs = sorted_qs.order_by('is_favorite', *current_ordering)
        
        return sorted_qs
    
    def get_context_data(self, **kwargs):
        from .models import FavoriteProduct
        
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        context['categories'] = Category.objects.filter(
            tenant=user.tenant, is_active=True
        )
        
        # Get favorite IDs for star rendering in template
        effective_location = self._get_user_location()
        if effective_location:
            context['favorite_ids'] = set(FavoriteProduct.objects.filter(
                tenant=user.tenant, location=effective_location
            ).values_list('product_id', flat=True))
        else:
            context['favorite_ids'] = set()
        
        # Determine user's location(s) for stock display
        user_location = user.location
        user_location_type = None
        
        # Map role to location type if user has no specific location
        if not user_location and user.role:
            role_location_map = {
                'PRODUCTION_MANAGER': 'PRODUCTION',
                'STORES_MANAGER': 'STORES',
                'SHOP_MANAGER': 'SHOP',
            }
            user_location_type = role_location_map.get(user.role.name)
        
        # Calculate stock by product for user's location
        stock_filter = {'product__tenant': user.tenant}
        if user_location:
            stock_filter['location'] = user_location
        elif user_location_type:
            stock_filter['location__location_type'] = user_location_type
        
        # Get stock quantities from ledger
        from django.db.models import Sum
        stock_data = InventoryLedger.objects.filter(
            **stock_filter
        ).values('product_id').annotate(
            total_stock=Sum('quantity')
        )
        
        # Create dict for easy lookup in template
        context['stock_by_product'] = {
            item['product_id']: item['total_stock'] or 0 
            for item in stock_data
        }
        
        # Pass location info for display
        if user_location:
            context['current_location'] = user_location
        elif user_location_type:
            context['current_location'] = Location.objects.filter(
                tenant=user.tenant,
                is_active=True,
                location_type=user_location_type
            ).first()
        else:
            context['current_location'] = None
        
        return context



class ToggleFavoriteView(LoginRequiredMixin, View):
    """Toggle favorite status of a product for the user's location."""
    
    def post(self, request, pk):
        from .models import FavoriteProduct
        
        product = get_object_or_404(Product, pk=pk, tenant=request.user.tenant)
        
        # Determine user's effective location
        location = request.user.location
        if not location and request.user.role:
            role_location_map = {
                'PRODUCTION_MANAGER': 'PRODUCTION',
                'STORES_MANAGER': 'STORES',
                'SHOP_MANAGER': 'SHOP',
            }
            loc_type = role_location_map.get(request.user.role.name)
            if loc_type:
                location = Location.objects.filter(
                    tenant=request.user.tenant, is_active=True, location_type=loc_type
                ).first()
        
        if not location:
            return JsonResponse({'error': 'No location assigned'}, status=400)
        
        # Toggle
        favorite, created = FavoriteProduct.objects.get_or_create(
            tenant=request.user.tenant,
            location=location,
            product=product,
            defaults={'created_by': request.user}
        )
        
        if not created:
            favorite.delete()
            return JsonResponse({'is_favorite': False})
        
        return JsonResponse({'is_favorite': True})


class ProductDetailView(LoginRequiredMixin, DetailView):
    """View product details including stock levels."""
    model = Product
    template_name = 'inventory/product_detail.html'
    context_object_name = 'product'
    
    def get_queryset(self):
        return Product.objects.filter(tenant=self.request.user.tenant)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        product = self.object
        
        # Get stock by location
        context['stock_by_location'] = InventoryLedger.objects.filter(
            product=product
        ).values(
            'location__name', 'location__location_type'
        ).annotate(
            total_stock=Sum('quantity')
        ).order_by('location__name')
        
        # Get active batches
        context['batches'] = Batch.objects.filter(
            product=product,
            status='AVAILABLE',
            current_quantity__gt=0
        ).select_related('location').order_by('expiry_date')
        
        # Get recent ledger entries
        context['recent_entries'] = InventoryLedger.objects.filter(
            product=product
        ).select_related('location', 'created_by')[:20]
        
        return context


class ProductCreateView(LoginRequiredMixin, CreateView):
    """Create a new product."""
    model = Product
    form_class = ProductForm
    template_name = 'inventory/product_form.html'
    success_url = reverse_lazy('inventory:product_list')
    
    def dispatch(self, request, *args, **kwargs):
        # Check if shop manager has permission
        if request.user.role and request.user.role.name == 'SHOP_MANAGER':
            if not request.user.tenant.shop_manager_can_add_products:
                messages.error(request, "You don't have permission to add products.")
                return redirect('inventory:product_list')
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['tenant'] = self.request.user.tenant
        return kwargs
    
    def form_valid(self, form):
        form.instance.tenant = self.request.user.tenant
        messages.success(self.request, f'Product "{form.instance.name}" created successfully!')
        return super().form_valid(form)


class ProductUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing product."""
    model = Product
    form_class = ProductForm
    template_name = 'inventory/product_form.html'
    success_url = reverse_lazy('inventory:product_list')
    
    def get_queryset(self):
        return Product.objects.filter(tenant=self.request.user.tenant)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['tenant'] = self.request.user.tenant
        return kwargs
    
    def form_valid(self, form):
        messages.success(self.request, f'Product "{form.instance.name}" updated!')
        return super().form_valid(form)


class ProductDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a product."""
    model = Product
    template_name = 'inventory/product_confirm_delete.html'
    success_url = reverse_lazy('inventory:product_list')
    
    def get_queryset(self):
        return Product.objects.filter(tenant=self.request.user.tenant)


class ProductTemplateDownloadView(LoginRequiredMixin, View):
    def get(self, request):
        # Restriction: PRODUCTION_MANAGER, STORES_MANAGER, ADMIN
        role_name = request.user.role.name if request.user.role else 'ATTENDANT'
        if role_name not in ['PRODUCTION_MANAGER', 'STORES_MANAGER', 'ADMIN']:
             messages.error(request, "Permission denied. Restricted to Managers and Admins.")
             return redirect('inventory:product_list')
             
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Product Template"
        
        # Headers
        headers = ['Name', 'Category', 'Description', 'Barcode', 'Unit', 'Cost Price', 'Selling Price', 'Alert Threshold']
        ws.append(headers)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=product_import_template.xlsx'
        
        wb.save(response)
        return response

class ProductBulkUploadView(LoginRequiredMixin, View):
    template_name = 'inventory/product_upload.html'
    
    def dispatch(self, request, *args, **kwargs):
        role_name = request.user.role.name if request.user.role else 'ATTENDANT'
        if role_name not in ['PRODUCTION_MANAGER', 'STORES_MANAGER', 'ADMIN']:
             messages.error(request, "Permission denied. Restricted to Managers and Admins.")
             return redirect('inventory:product_list')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        return render(request, self.template_name)
    
    def post(self, request):
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Please select a file to upload.")
            return redirect('inventory:product_upload')
            
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            created_count = 0
            errors = []
            
            # Skip header
            rows = list(ws.rows)
            if len(rows) < 2:
                 messages.warning(request, "File contains no data.")
                 return redirect('inventory:product_upload')
                 
            # Skip header and instructions rows (rows 1 and 2)
            for index, row in enumerate(rows[2:], start=3):
                try:
                    # Read values
                    name = row[0].value
                    category_name = row[1].value
                    description = row[2].value or ""
                    sku = str(row[3].value).strip() if row[3].value else None
                    unit = row[4].value
                    selling_price = row[5].value
                    alert_threshold = row[6].value
                    
                    # Skip completely empty rows
                    if not any([name, category_name, sku, unit, selling_price]):
                        continue
                    
                    # Validate required fields
                    if not name:
                        errors.append(f"Row {index}: Product name is required")
                        continue
                    
                    # Auto-generate SKU if not provided
                    if not sku:
                        sku = f"SKU-{request.user.tenant.id}-{index}"
                    
                    # Validate and default unit
                    if not unit:
                        unit = "UNIT"
                    else:
                        unit = unit.upper().strip()
                        valid_units = [choice[0] for choice in Product.UNIT_CHOICES]
                        if unit not in valid_units:
                            errors.append(f"Row {index}: Invalid unit '{unit}'. Valid options: {', '.join(valid_units)}")
                            continue
                    
                    # Check duplicate SKU
                    if Product.objects.filter(tenant=request.user.tenant, sku=sku).exists():
                        errors.append(f"Row {index}: Duplicate SKU '{sku}' - this SKU already exists")
                        continue
                    
                    # Handle category
                    category = None
                    if category_name:
                        category_name = str(category_name).strip()
                        category, _ = Category.objects.get_or_create(
                            tenant=request.user.tenant,
                            name__iexact=category_name,
                            defaults={'name': category_name}
                        )
                    
                    # Validate and convert numeric fields
                    try:
                        reorder_level = Decimal(str(alert_threshold)) if alert_threshold else Decimal('0')
                        if reorder_level < 0:
                            raise ValueError("Alert threshold cannot be negative")
                    except (ValueError, decimal.InvalidOperation) as e:
                        errors.append(f"Row {index}: Invalid alert threshold '{alert_threshold}' - must be a positive number")
                        continue
                    
                    default_price = Decimal('0')
                    if selling_price:
                        try:
                            default_price = Decimal(str(selling_price))
                            if default_price < 0:
                                raise ValueError("Price cannot be negative")
                        except (ValueError, decimal.InvalidOperation):
                            errors.append(f"Row {index}: Invalid selling price '{selling_price}' - must be a positive number")
                            continue
                    
                    # Create Product
                    product = Product.objects.create(
                        tenant=request.user.tenant,
                        name=name.strip(),
                        category=category,
                        description=description.strip() if description else "",
                        sku=sku,
                        unit_of_measure=unit,
                        reorder_level=reorder_level,
                        default_selling_price=default_price
                    )
                    
                    # Create Shop Price only if selling price is provided and user has shop location
                    if selling_price and request.user.location and request.user.location.location_type == 'SHOP':
                        try:
                            ShopPrice.objects.create(
                                tenant=request.user.tenant,
                                location=request.user.location,
                                product=product,
                                selling_price=default_price
                            )
                        except Exception as e:
                            # Product created but shop price failed - log but don't fail
                            errors.append(f"Row {index}: Product created but shop price failed: {str(e)}")
                    
                    created_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {index}: Unexpected error - {str(e)}")
            
            # Show results
            if created_count > 0 and not errors:
                messages.success(request, f"✅ Successfully imported {created_count} product(s)!")
            elif created_count > 0 and errors:
                messages.warning(request, f"⚠️ Partially successful: {created_count} product(s) created, {len(errors)} error(s)")
                # Show first 10 errors
                for error in errors[:10]:
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.info(request, f"... and {len(errors) - 10} more error(s)")
            elif errors:
                messages.error(request, f"❌ Import failed: {len(errors)} error(s) found. No products were created.")
                for error in errors[:10]:
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.info(request, f"... and {len(errors) - 10} more error(s)")
            else:
                messages.warning(request, "No valid data found in the file.")
                
            return redirect('inventory:product_list')
            
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            return redirect('inventory:product_upload')


class InventoryExportView(LoginRequiredMixin, View):
    """
    Export inventory to Excel.
    Allowed: STORES_MANAGER, SHOP_MANAGER, ADMIN
    """
    def get(self, request):
        role_name = request.user.role.name if request.user.role else ''
        if role_name not in ['STORES_MANAGER', 'SHOP_MANAGER', 'ADMIN']:
            messages.error(request, "Permission denied.")
            return redirect('inventory:product_list')
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory"
        
        # Headers
        headers = ['Name', 'Category', 'SKU', 'Unit', 'Stock Qty', 'Selling Price']
        # If Admin or Stores Manager, maybe include Cost Price?
        if role_name in ['ADMIN', 'STORES_MANAGER']:
            headers.append('Cost Price')
            
        ws.append(headers)
        
        # Filter products based on location
        products = Product.objects.filter(tenant=request.user.tenant, is_active=True).select_related('category')
        
        location = request.user.location
        
        for product in products:
            qty = 0
            price = 0
            cost = 0 
            
            # If user has a location, get stock/price for that location
            if location:
                # Stock (Need to implement Ledger aggregation later or use existing logic)
                stock_in = InventoryLedger.objects.filter(product=product, location=location, quantity__gt=0).aggregate(Sum('quantity'))['quantity__sum'] or 0
                stock_out = InventoryLedger.objects.filter(product=product, location=location, quantity__lt=0).aggregate(Sum('quantity'))['quantity__sum'] or 0
                qty = stock_in + stock_out
                
                # Price
                shop_price = ShopPrice.objects.filter(product=product, location=location).first()
                if shop_price:
                    price = shop_price.selling_price
                
                # Cost - get average from available batches
                avg_cost = Batch.objects.filter(
                    product=product,
                    location=location,
                    status='AVAILABLE',
                    current_quantity__gt=0
                ).aggregate(Avg('unit_cost'))['unit_cost__avg']
                cost = avg_cost if avg_cost else Decimal('0')
            else:
                # Admin without location
                pass

            row = [
                product.name,
                product.category.name if product.category else '',
                product.sku or '',
                product.unit_of_measure,
                qty,
                price
            ]
            
            if role_name in ['ADMIN', 'STORES_MANAGER']:
                row.append(cost)
                
            ws.append(row)
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=inventory_export.xlsx'
        
        wb.save(response)
        return response


# ============ Batch Views ============
from apps.core.mixins import SortableMixin

class BatchListView(LoginRequiredMixin, SortableMixin, ListView):
    """List all batches for the tenant."""
    model = Batch
    template_name = 'inventory/batch_list.html'
    context_object_name = 'batches'
    sortable_fields = ['created_at', 'product__name', 'batch_number', 'unit_cost', 'initial_quantity', 'current_quantity', 'expiry_date', 'status', 'location__name']
    default_sort = '-created_at'
    
    def get_queryset(self):
        queryset = Batch.objects.filter(
            tenant=self.request.user.tenant
        ).select_related('product', 'location')
        
        # Status filter
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Location filter
        location = self.request.GET.get('location')
        if location:
            queryset = queryset.filter(location_id=location)
        
        
        return self.apply_sorting(queryset)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['locations'] = Location.objects.filter(
            tenant=self.request.user.tenant,
            is_active=True
        )
        return context


class BatchCreateView(LoginRequiredMixin, CreateView):
    """Create a new batch (receive stock)."""
    model = Batch
    form_class = BatchForm
    template_name = 'inventory/batch_form.html'
    success_url = reverse_lazy('inventory:batch_list')
    
    def dispatch(self, request, *args, **kwargs):
        # Check if shop manager has permission
        if request.user.role and request.user.role.name == 'SHOP_MANAGER':
            if not request.user.tenant.shop_manager_can_receive_stock:
                messages.error(request, "You don't have permission to receive stock.")
                return redirect('inventory:batch_list')
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['tenant'] = self.request.user.tenant
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = context.get('form')
        if form and hasattr(form, 'auto_location'):
            context['auto_location'] = form.auto_location
        return context
    
    def form_valid(self, form):
        form.instance.tenant = self.request.user.tenant
        # Initialize current_quantity to 0. The ledger entry creation below will 
        # increment this by the received quantity via the InventoryLedger.save() method.
        # This prevents double counting (once from here, once from ledger signal).
        form.instance.current_quantity = Decimal('0')
        
        # Save the batch first
        response = super().form_valid(form)
        
        # Create ledger entry for stock in
        InventoryLedger.objects.create(
            tenant=self.request.user.tenant,
            product=self.object.product,
            batch=self.object,
            location=self.object.location,
            transaction_type='IN',
            quantity=self.object.initial_quantity,
            unit_cost=self.object.unit_cost,
            reference_type='Batch',
            reference_id=self.object.pk,
            notes=f"Initial stock receipt - Batch {self.object.batch_number}",
            created_by=self.request.user
        )
        
        messages.success(self.request, f'Batch "{self.object.batch_number}" received successfully!')
        return response


class BatchDetailView(LoginRequiredMixin, DetailView):
    """View batch details."""
    model = Batch
    template_name = 'inventory/batch_detail.html'
    context_object_name = 'batch'
    
    def get_queryset(self):
        return Batch.objects.filter(tenant=self.request.user.tenant)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['ledger_entries'] = InventoryLedger.objects.filter(
            batch=self.object
        ).select_related('created_by').order_by('-created_at')
        return context


# ============ Stock Views ============
class StockOverviewView(LoginRequiredMixin, View):
    """Overview of stock across all locations."""
    template_name = 'inventory/stock_overview.html'
    
    def get(self, request):
        from django.core.paginator import Paginator
        from .models import FavoriteProduct
        
        tenant = request.user.tenant
        user = request.user
        role_name = user.role.name if user.role else ''
        is_shop_role = role_name in ('SHOP_MANAGER', 'SHOP_ATTENDANT')
        
        # Determine if the user should see stock by location
        show_stock_by_location = True
        if is_shop_role and not tenant.shops_can_see_other_stock:
            show_stock_by_location = False
        
        # Stock by product and location (only if allowed)
        stock_summary = None
        if show_stock_by_location:
            stock_summary_qs = InventoryLedger.objects.filter(
                tenant=tenant
            ).values(
                'product__id', 'product__name', 'product__sku',
                'location__id', 'location__name', 'location__location_type'
            ).annotate(
                total_stock=Sum('quantity')
            ).filter(total_stock__gt=0).order_by('product__name', 'location__name')
            
            # Pagination for Stock Summary
            page_summary = request.GET.get('page_summary', 1)
            paginator_summary = Paginator(stock_summary_qs, 25)
            stock_summary = paginator_summary.get_page(page_summary)
        
        # Determine user's effective location for favorites
        effective_location = user.location
        if not effective_location and user.role:
            role_location_map = {
                'PRODUCTION_MANAGER': 'PRODUCTION',
                'STORES_MANAGER': 'STORES',
                'SHOP_MANAGER': 'SHOP',
            }
            loc_type = role_location_map.get(role_name)
            if loc_type:
                effective_location = Location.objects.filter(
                    tenant=tenant, is_active=True, location_type=loc_type
                ).first()
        
        # Get favorite product IDs for this location
        favorite_product_ids = set()
        if effective_location:
            favorite_product_ids = set(FavoriteProduct.objects.filter(
                tenant=tenant, location=effective_location
            ).values_list('product_id', flat=True))
        
        # Low stock alerts
        low_stock_list = []
        products = Product.objects.filter(tenant=tenant, is_active=True)
        for product in products:
            total = product.get_total_stock()
            if total <= product.reorder_level:
                low_stock_list.append({
                    'product': product,
                    'current_stock': total,
                    'reorder_level': product.reorder_level,
                    'is_favorite': product.pk in favorite_product_ids,
                })
        
        # Sort low stock: favorites first, then by current stock ascending
        low_stock_list.sort(key=lambda x: (0 if x['is_favorite'] else 1, x['current_stock']))
        
        # Expiring soon (within 30 days)
        from django.utils import timezone
        from datetime import timedelta
        expiring_batches = Batch.objects.filter(
            tenant=tenant,
            status='AVAILABLE',
            current_quantity__gt=0,
            expiry_date__lte=timezone.now().date() + timedelta(days=30),
            expiry_date__gte=timezone.now().date()
        ).select_related('product', 'location').order_by('expiry_date')[:10]
        
        # Pagination for Low Stock
        page_low_stock = request.GET.get('page_low_stock', 1)
        paginator_low_stock = Paginator(low_stock_list, 10)
        low_stock = paginator_low_stock.get_page(page_low_stock)
        
        context = {
            'stock_summary': stock_summary,
            'low_stock': low_stock,
            'expiring_batches': expiring_batches,
            'show_stock_by_location': show_stock_by_location,
            'favorite_product_ids': favorite_product_ids,
        }
        
        return render(request, self.template_name, context)


class StockAdjustmentView(LoginRequiredMixin, View):
    """Create stock adjustments with role-based restrictions."""
    template_name = 'inventory/stock_adjustment.html'
    
    def get(self, request):
        form = StockAdjustmentForm(tenant=request.user.tenant, user=request.user)
        return render(request, self.template_name, {'form': form})
    
    def post(self, request):
        form = StockAdjustmentForm(request.POST, tenant=request.user.tenant, user=request.user)
        if form.is_valid():
            product = form.cleaned_data['product']
            location = form.cleaned_data['location']
            batch = form.cleaned_data.get('batch')
            quantity = form.cleaned_data['quantity']
            adjustment_type = form.cleaned_data['adjustment_type']
            reason = form.cleaned_data['reason']
            
            user = request.user
            role_name = user.role.name if hasattr(user, 'role') and user.role else ''
            is_own_location = (user.location_id == location.pk)
            
            # Determine if this needs approval
            needs_approval = False
            if role_name == 'STORES_MANAGER' and not is_own_location:
                needs_approval = True
            
            # Admin always auto-approves
            if role_name == 'ADMIN':
                needs_approval = False
            
            # Create the StockAdjustment record
            adjustment = StockAdjustment.objects.create(
                tenant=user.tenant,
                product=product,
                batch=batch,
                location=location,
                adjustment_type=adjustment_type,
                quantity=quantity,
                reason=reason,
                requested_by=user,
                status='PENDING' if needs_approval else 'APPROVED',
            )
            
            if needs_approval:
                # Notify the target location's manager(s) and admins
                from apps.notifications.models import Notification
                from apps.core.models import User
                
                # Find managers of the target location + admins
                target_users = User.objects.filter(
                    tenant=user.tenant,
                    is_active=True,
                ).filter(
                    Q(location=location, role__name='SHOP_MANAGER') |
                    Q(role__name='ADMIN')
                ).exclude(pk=user.pk)
                
                for target_user in target_users:
                    Notification.objects.create(
                        tenant=user.tenant,
                        user=target_user,
                        title='Stock Adjustment Pending Approval',
                        message=f'{user.get_full_name() or user.email} has requested a stock adjustment of {quantity} '
                                f'for {product.name} at {location.name}. Reason: {reason}',
                        notification_type='STOCK_ADJUSTMENT',
                        reference_type='StockAdjustment',
                        reference_id=adjustment.pk,
                    )
                
                messages.info(request, 
                    f'Stock adjustment for {product.name} at {location.name} has been submitted for approval.'
                )
            else:
                # Auto-approve: create ledger entry immediately
                adjustment.approve(user, notes='Auto-approved (own location or admin)')
                messages.success(request, f'Stock adjustment recorded for {product.name}.')
            
            return redirect('inventory:adjustment_history')
        
        return render(request, self.template_name, {'form': form})


class AdjustmentHistoryView(LoginRequiredMixin, SortableMixin, ListView):
    """Standalone adjustment history showing all adjustments."""
    model = StockAdjustment
    template_name = 'inventory/adjustment_history.html'
    context_object_name = 'adjustments'
    sortable_fields = ['created_at', 'product__name', 'quantity', 'adjustment_type', 'location__name', 'status']
    default_sort = '-created_at'
    paginate_by = 25
    
    def get_queryset(self):
        user = self.request.user
        queryset = StockAdjustment.objects.filter(
            tenant=user.tenant
        ).select_related(
            'product', 'location', 'batch', 'requested_by', 'reviewed_by'
        )
        
        # Role-based filtering
        role_name = user.role.name if hasattr(user, 'role') and user.role else ''
        if role_name in ('SHOP_MANAGER', 'SHOP_ATTENDANT'):
            # See own location adjustments only
            queryset = queryset.filter(location=user.location)
        elif role_name == 'PRODUCTION_MANAGER':
            queryset = queryset.filter(location=user.location)
        # Stores Manager, Admin, Auditor, Accountant: see all
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filter by location
        location_id = self.request.GET.get('location')
        if location_id:
            queryset = queryset.filter(location_id=location_id)
        
        # Search
        search = self.request.GET.get('q')
        if search:
            queryset = queryset.filter(
                Q(product__name__icontains=search) |
                Q(reason__icontains=search) |
                Q(requested_by__first_name__icontains=search) |
                Q(requested_by__last_name__icontains=search)
            )
        
        return self.apply_sorting(queryset)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        role_name = user.role.name if hasattr(user, 'role') and user.role else ''
        
        context['current_status'] = self.request.GET.get('status', '')
        context['current_location'] = self.request.GET.get('location', '')
        context['current_search'] = self.request.GET.get('q', '')
        context['status_choices'] = StockAdjustment.STATUS_CHOICES
        
        # Pending count for the user's location (for the badge)
        if role_name in ('SHOP_MANAGER', 'ADMIN'):
            if role_name == 'SHOP_MANAGER' and user.location:
                context['pending_count'] = StockAdjustment.objects.filter(
                    tenant=user.tenant,
                    location=user.location,
                    status='PENDING'
                ).count()
            elif role_name == 'ADMIN':
                context['pending_count'] = StockAdjustment.objects.filter(
                    tenant=user.tenant,
                    status='PENDING'
                ).count()
        
        # Locations for filter dropdown
        if role_name in ('STORES_MANAGER', 'ADMIN', 'AUDITOR', 'ACCOUNTANT'):
            context['locations'] = Location.objects.filter(
                tenant=user.tenant, is_active=True
            )
        
        return context


class ReviewAdjustmentView(LoginRequiredMixin, View):
    """Approve or reject a pending stock adjustment."""
    
    def post(self, request, pk):
        adjustment = get_object_or_404(
            StockAdjustment,
            pk=pk,
            tenant=request.user.tenant,
            status='PENDING'
        )
        
        user = request.user
        role_name = user.role.name if hasattr(user, 'role') and user.role else ''
        
        # Only Shop Manager of that location or Admin can review
        can_review = False
        if role_name == 'ADMIN':
            can_review = True
        elif role_name == 'SHOP_MANAGER' and user.location_id == adjustment.location_id:
            can_review = True
        
        if not can_review:
            messages.error(request, 'You do not have permission to review this adjustment.')
            return redirect('inventory:adjustment_history')
        
        action = request.POST.get('action')
        review_notes = request.POST.get('review_notes', '')
        
        from apps.notifications.models import Notification
        
        if action == 'approve':
            adjustment.approve(user, notes=review_notes)
            messages.success(request, 
                f'Adjustment for {adjustment.product.name} at {adjustment.location.name} has been approved.'
            )
            # Notify the requester
            if adjustment.requested_by and adjustment.requested_by != user:
                Notification.objects.create(
                    tenant=user.tenant,
                    user=adjustment.requested_by,
                    title='Stock Adjustment Approved',
                    message=f'Your stock adjustment of {adjustment.quantity} for {adjustment.product.name} '
                            f'at {adjustment.location.name} has been approved by {user.get_full_name() or user.email}.',
                    notification_type='STOCK_ADJUSTMENT',
                    reference_type='StockAdjustment',
                    reference_id=adjustment.pk,
                )
        elif action == 'reject':
            adjustment.reject(user, notes=review_notes)
            messages.warning(request, 
                f'Adjustment for {adjustment.product.name} at {adjustment.location.name} has been rejected.'
            )
            # Notify the requester
            if adjustment.requested_by and adjustment.requested_by != user:
                Notification.objects.create(
                    tenant=user.tenant,
                    user=adjustment.requested_by,
                    title='Stock Adjustment Rejected',
                    message=f'Your stock adjustment of {adjustment.quantity} for {adjustment.product.name} '
                            f'at {adjustment.location.name} has been rejected by {user.get_full_name() or user.email}.'
                            f'{" Reason: " + review_notes if review_notes else ""}',
                    notification_type='STOCK_ADJUSTMENT',
                    reference_type='StockAdjustment',
                    reference_id=adjustment.pk,
                )
        
        return redirect('inventory:adjustment_history')


class InventoryLedgerListView(LoginRequiredMixin, SortableMixin, ListView):
    """
    Complete audit trail of all inventory movements.
    Read-only view for Auditors and Admins.
    """
    model = InventoryLedger
    template_name = 'inventory/inventory_ledger.html'
    context_object_name = 'entries'
    sortable_fields = ['created_at', 'product__name', 'quantity', 'unit_cost', 'transaction_type', 'location__name', 'batch__batch_number']
    default_sort = '-created_at'
    
    def get_queryset(self):
        from django.utils import timezone
        from datetime import timedelta
        
        user = self.request.user
        queryset = InventoryLedger.objects.filter(
            tenant=user.tenant
        ).select_related(
            'product', 'location', 'batch', 'created_by'
        ).order_by('-created_at')
        
        # Date range filter
        date_range = self.request.GET.get('range', 'all')
        today = timezone.now().date()
        
        if date_range == 'today':
            queryset = queryset.filter(created_at__date=today)
        elif date_range == 'week':
            queryset = queryset.filter(created_at__date__gte=today - timedelta(days=7))
        elif date_range == 'month':
            queryset = queryset.filter(created_at__date__gte=today - timedelta(days=30))
        
        # Transaction type filter
        tx_type = self.request.GET.get('type')
        if tx_type:
            queryset = queryset.filter(transaction_type=tx_type)
        
        # Location filter
        location = self.request.GET.get('location')
        if location:
            queryset = queryset.filter(location_id=location)
        
        # Product search
        product = self.request.GET.get('product')
        if product:
            queryset = queryset.filter(
                Q(product__name__icontains=product) |
                Q(product__sku__icontains=product)
            )
        
        return self.apply_sorting(queryset)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['transaction_types'] = InventoryLedger.TRANSACTION_TYPES
        context['locations'] = Location.objects.filter(
            tenant=self.request.user.tenant,
            is_active=True
        )
        context['current_range'] = self.request.GET.get('range', 'all')
        context['current_type'] = self.request.GET.get('type', '')
        context['current_location'] = self.request.GET.get('location', '')
        context['current_product'] = self.request.GET.get('product', '')
        return context


# ============ API Views ============
def get_batches_for_product(request):
    """AJAX endpoint to get batches for a product at a location."""
    product_id = request.GET.get('product_id')
    location_id = request.GET.get('location_id')
    
    if not product_id or not location_id:
        return JsonResponse({'batches': []})
    
    batches = Batch.objects.filter(
        product_id=product_id,
        location_id=location_id,
        status='AVAILABLE',
        current_quantity__gt=0,
        tenant=request.user.tenant
    ).values('id', 'batch_number', 'current_quantity', 'unit_cost', 'expiry_date')
    
    return JsonResponse({'batches': list(batches)})


def search_products(request):
    """AJAX endpoint for product search autocomplete."""
    query = request.GET.get('q', '')
    
    if not query or len(query) < 2:
        return JsonResponse({'products': []})
    
    products = Product.objects.filter(
        tenant=request.user.tenant,
        is_active=True
    ).filter(
        Q(name__icontains=query) | 
        Q(sku__icontains=query)
    )[:10]  # Limit to 10 suggestions
    
    results = [
        {
            'id': p.pk,
            'name': p.name,
            'sku': p.sku,
            'category': p.category.name if p.category else None,
        }
        for p in products
    ]
    
    return JsonResponse({'products': results})


# ============ Shop Price Management ============

class ShopPriceListView(LoginRequiredMixin, View):
    """
    List all products with their shop price status for the shop manager's location.
    """
    template_name = 'inventory/shop_price_list.html'
    
    def dispatch(self, request, *args, **kwargs):
        role_name = request.user.role.name if request.user.role else None
        if role_name not in ['SHOP_MANAGER', 'ADMIN']:
            messages.error(request, 'Only shop managers can access pricing.')
            return redirect('core:dashboard')
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request):
        user = request.user
        shop = user.location
        
        # Admin can select a shop
        if not shop or shop.location_type != 'SHOP':
            shop_id = request.GET.get('shop')
            if shop_id:
                shop = Location.objects.filter(
                    tenant=user.tenant, pk=shop_id, location_type='SHOP'
                ).first()
        
        context = {'shop': shop}
        
        if not shop or shop.location_type != 'SHOP':
            context['shops'] = Location.objects.filter(
                tenant=user.tenant, location_type='SHOP', is_active=True
            )
            return render(request, self.template_name, context)
        
        # Get all active products
        products = Product.objects.filter(
            tenant=user.tenant,
            is_active=True
        ).select_related('category').prefetch_related('shop_prices')
        
        # Build product list with price status
        products_with_status = []
        for product in products:
            shop_price = product.shop_prices.filter(
                location=shop, is_active=True
            ).first()
            
            products_with_status.append({
                'product': product,
                'shop_price': shop_price,
                'has_price': shop_price is not None,
                'selling_price': shop_price.selling_price if shop_price else None,
            })
        
        # Filter by status
        status_filter = request.GET.get('status')
        if status_filter == 'with_price':
            products_with_status = [p for p in products_with_status if p['has_price']]
        elif status_filter == 'without_price':
            products_with_status = [p for p in products_with_status if not p['has_price']]
        
        # Search filter
        search = request.GET.get('q', '').lower()
        if search:
            products_with_status = [
                p for p in products_with_status 
                if search in p['product'].name.lower() or search in (p['product'].sku or '').lower()
            ]
        
        context['products'] = products_with_status
        context['total_products'] = len(products_with_status)
        context['priced_count'] = sum(1 for p in products_with_status if p['has_price'])
        
        return render(request, self.template_name, context)


class ShopPriceSetView(LoginRequiredMixin, View):
    """
    Set or update shop price for a product.
    """
    template_name = 'inventory/shop_price_form.html'
    
    def dispatch(self, request, *args, **kwargs):
        role_name = request.user.role.name if request.user.role else None
        if role_name not in ['SHOP_MANAGER', 'ADMIN']:
            messages.error(request, 'Only shop managers can set pricing.')
            return redirect('core:dashboard')
        return super().dispatch(request, *args, **kwargs)
    
    def get_shop_and_product(self, request, product_pk):
        user = request.user
        shop = user.location
        
        if not shop or shop.location_type != 'SHOP':
            shop_id = request.GET.get('shop')
            if shop_id:
                shop = Location.objects.filter(
                    tenant=user.tenant, pk=shop_id, location_type='SHOP'
                ).first()
        
        product = get_object_or_404(Product, pk=product_pk, tenant=user.tenant)
        return shop, product
    
    def get(self, request, pk):
        shop, product = self.get_shop_and_product(request, pk)
        
        if not shop:
            messages.error(request, 'No shop selected.')
            return redirect('inventory:shop_price_list')
        
        # Get existing price or create new
        existing_price = ShopPrice.objects.filter(
            product=product,
            location=shop,
            is_active=True
        ).first()
        
        initial = {'selling_price': product.default_selling_price}
        if existing_price:
            initial['selling_price'] = existing_price.selling_price
        
        context = {
            'product': product,
            'shop': shop,
            'existing_price': existing_price,
            'initial_price': initial['selling_price'],
        }
        
        return render(request, self.template_name, context)
    
    def post(self, request, pk):
        shop, product = self.get_shop_and_product(request, pk)
        
        if not shop:
            messages.error(request, 'No shop selected.')
            return redirect('inventory:shop_price_list')
        
        selling_price = request.POST.get('selling_price', '0')
        
        try:
            selling_price = Decimal(selling_price)
            if selling_price <= 0:
                raise ValueError("Price must be positive")
        except:
            messages.error(request, 'Invalid price. Please enter a valid number.')
            return redirect('inventory:shop_price_set', pk=pk)
        
        # Deactivate old prices for this product/shop
        ShopPrice.objects.filter(
            product=product,
            location=shop,
            is_active=True
        ).update(is_active=False)
        
        # Create new shop price
        ShopPrice.objects.create(
            tenant=request.user.tenant,
            product=product,
            location=shop,
            selling_price=selling_price,
            is_active=True
        )
        
        messages.success(request, f'Price for "{product.name}" set to {request.user.tenant.currency_symbol}{selling_price}')
        return redirect('inventory:shop_price_list')


class ProductTemplateDownloadView(LoginRequiredMixin, View):
    """Download Excel template for bulk product upload."""
    
    def get(self, request):
        from openpyxl.worksheet.datavalidation import DataValidation
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Headers
        headers = ['Name*', 'Category', 'Description', 'SKU', 'Unit', 'Selling Price', 'Alert Threshold']
        ws.append(headers)
        
        # Instructions row
        ws.append([
            'Enter product name (required)',
            'Select from dropdown or leave blank',
            'Optional description',
            'Leave blank to auto-generate',
            'Select from dropdown',
            'Optional (can be set later)',
            'Optional stock alert level'
        ])
        
        # Sample data row
        ws.append([
            'Sample Product',
            '',  # Will be dropdown
            'This is a sample product',
            '',  # Auto-generated
            'UNIT',  # Will be dropdown
            '100.00',
            '10'
        ])
        
        # Get categories for dropdown
        categories = list(Category.objects.filter(
            tenant=request.user.tenant
        ).values_list('name', flat=True))
        
        # Create dropdown for Category column (B3:B1000)
        if categories:
            category_dv = DataValidation(
                type="list",
                formula1=f'"{",".join(categories)}"',
                allow_blank=True
            )
            category_dv.error = 'Please select a category from the list or leave blank'
            category_dv.errorTitle = 'Invalid Category'
            ws.add_data_validation(category_dv)
            category_dv.add('B3:B1000')
        
        # Create dropdown for Unit column (E3:E1000)
        units = [choice[0] for choice in Product.UNIT_CHOICES]
        unit_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(units)}"',
            allow_blank=False
        )
        unit_dv.error = 'Please select a valid unit from the list'
        unit_dv.errorTitle = 'Invalid Unit'
        unit_dv.prompt = 'Select unit of measure'
        unit_dv.promptTitle = 'Unit Selection'
        ws.add_data_validation(unit_dv)
        unit_dv.add('E3:E1000')
        
        # Style the header row
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 25  # Name
        ws.column_dimensions['B'].width = 20  # Category
        ws.column_dimensions['C'].width = 30  # Description
        ws.column_dimensions['D'].width = 15  # SKU
        ws.column_dimensions['E'].width = 12  # Unit
        ws.column_dimensions['F'].width = 15  # Selling Price
        ws.column_dimensions['G'].width = 15  # Alert Threshold
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=product_import_template.xlsx'
        
        wb.save(response)
        return response


@login_required
def get_adjustment_details_api(request, pk):
    """AJAX endpoint to get stock adjustment details for review modal."""
    from .models import StockAdjustment
    
    adjustment = get_object_or_404(
        StockAdjustment, 
        pk=pk, 
        tenant=request.user.tenant
    )
    
    # Check if the user has permission to review this adjustment
    role_name = request.user.role.name if hasattr(request.user, 'role') and request.user.role else ''
    can_review = False
    if role_name == 'ADMIN':
        can_review = True
    elif role_name == 'SHOP_MANAGER' and request.user.location_id == adjustment.location_id:
        can_review = True
    
    data = {
        'id': adjustment.id,
        'product_name': adjustment.product.name,
        'quantity': float(adjustment.quantity),
        'location_name': adjustment.location.name,
        'reason': adjustment.get_adjustment_type_display(),
        'notes': adjustment.reason,
        'status': adjustment.status,
        'requested_by': adjustment.requested_by.get_full_name() or adjustment.requested_by.email if adjustment.requested_by else 'System',
        'created_at': adjustment.created_at.strftime('%Y-%m-%d %H:%M'),
        'can_review': can_review and adjustment.status == 'PENDING',
    }
    
    return JsonResponse(data)


# ============ Goods Receipt Views ============
from django.db import transaction
from django.urls import reverse
from .models import GoodsReceipt, GoodsReceiptItem
from .forms import GoodsReceiptForm, GoodsReceiptItemFormSet

class GoodsReceiptListView(LoginRequiredMixin, ListView):
    """List all goods receipts for the tenant."""
    model = GoodsReceipt
    template_name = 'inventory/goods_receipt_list.html'
    context_object_name = 'receipts'

    def get_queryset(self):
        qs = GoodsReceipt.objects.filter(tenant=self.request.user.tenant).order_by('-created_at')
        if getattr(self.request.user, 'role', None) and self.request.user.role.name == 'SHOP_MANAGER':
            qs = qs.filter(location=self.request.user.location)
        return qs

class GoodsReceiptCreateView(LoginRequiredMixin, CreateView):
    model = GoodsReceipt
    form_class = GoodsReceiptForm
    template_name = 'inventory/goods_receipt_form.html'
    success_url = reverse_lazy('inventory:goods_receipt_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['tenant'] = self.request.user.tenant
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['items'] = GoodsReceiptItemFormSet(self.request.POST, tenant=self.request.user.tenant)
        else:
            data['items'] = GoodsReceiptItemFormSet(tenant=self.request.user.tenant)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        items = context['items']
        with transaction.atomic():
            form.instance.tenant = self.request.user.tenant
            form.instance.created_by = self.request.user
            self.object = form.save()
            if items.is_valid():
                items.instance = self.object
                # Set tenant on each unsaved item before saving the formset
                for item_form in items:
                    if item_form.cleaned_data and not item_form.cleaned_data.get('DELETE', False):
                        item_form.instance.tenant = self.request.user.tenant
                items.save()
                messages.success(self.request, f"Goods Receipt created successfully (Status: Pending Verification).")
                return redirect('inventory:goods_receipt_list')
            else:
                return self.render_to_response(self.get_context_data(form=form))

class GoodsReceiptDetailView(LoginRequiredMixin, DetailView):
    model = GoodsReceipt
    template_name = 'inventory/goods_receipt_detail.html'
    context_object_name = 'receipt'

    def get_queryset(self):
        return GoodsReceipt.objects.filter(tenant=self.request.user.tenant)

@login_required
def verify_goods_receipt(request, pk):
    receipt = get_object_or_404(GoodsReceipt, pk=pk, tenant=request.user.tenant)
    if receipt.status != 'PENDING':
        messages.error(request, "Receipt is already verified or cancelled.")
        return redirect('inventory:goods_receipt_detail', pk=pk)

    # Check restriction
    if request.user.tenant.require_accountant_for_bulk_receiving:
        # User must be an ACCOUNTANT or ADMIN
        if request.user.role and request.user.role.name not in ['ACCOUNTANT', 'ADMIN']:
            messages.error(request, "Verification requires an Accountant or Admin.")
            return redirect('inventory:goods_receipt_detail', pk=pk)
    else:
        # Shop manager is enough
        if request.user.role and request.user.role.name not in ['SHOP_MANAGER', 'ADMIN', 'STORES_MANAGER']:
            messages.error(request, "You do not have permission to verify.")
            return redirect('inventory:goods_receipt_detail', pk=pk)

    if request.method == 'POST':
        try:
            with transaction.atomic():
                receipt.verify(verified_by=request.user)
            messages.success(request, f"Goods Receipt #{receipt.id} verified and stock updated.")
        except Exception as e:
            messages.error(request, f"Verification failed: {str(e)}")
            
    return redirect('inventory:goods_receipt_detail', pk=pk)


@login_required
def api_product_autocomplete(request):
    """Return JSON list of products matching query string for autocomplete inputs."""
    query = request.GET.get('q', '').strip()
    if not query or len(query) < 2:
        return JsonResponse({'results': []})

    products = Product.objects.filter(
        tenant=request.user.tenant,
        is_active=True
    ).filter(
        Q(name__icontains=query) | Q(sku__icontains=query)
    ).values('id', 'name', 'sku', 'unit_of_measure')[:15]

    return JsonResponse({'results': list(products)})
