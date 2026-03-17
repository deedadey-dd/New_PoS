"""
Shared Excel export utilities.
Provides helpers for creating styled Excel workbooks and HTTP responses.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)
ALT_ROW_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")


def create_export_workbook(title, headers, rows):
    """
    Create a styled Excel workbook with one sheet.

    Args:
        title: Sheet title (max 31 chars for Excel compatibility)
        headers: List of column header strings
        rows: List of lists/tuples, one per data row

    Returns:
        openpyxl.Workbook
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    _write_sheet(ws, headers, rows)
    return wb


def add_sheet(workbook, title, headers, rows):
    """
    Add an additional styled sheet to an existing workbook.

    Args:
        workbook: openpyxl.Workbook
        title: Sheet title
        headers: List of column header strings
        rows: List of lists/tuples

    Returns:
        The new worksheet
    """
    ws = workbook.create_sheet(title=title[:31])
    _write_sheet(ws, headers, rows)
    return ws


def _write_sheet(ws, headers, rows):
    """Write headers and rows to a worksheet with styling."""
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            # Alternate row shading
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row_idx in range(2, len(rows) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        # Cap width at 50, min at header length
        adjusted_width = min(max_length + 3, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"


def build_excel_response(workbook, filename):
    """
    Wrap an openpyxl Workbook in an HttpResponse for download.

    Args:
        workbook: openpyxl.Workbook
        filename: Download filename (should end with .xlsx)

    Returns:
        django.http.HttpResponse
    """
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response
